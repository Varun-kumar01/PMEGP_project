#!/usr/bin/env python3
"""Verify 10-sheet file and apply correct 68-column headers"""

from openpyxl import load_workbook

file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_10_SHEETS_CLEANED.xlsx'

CORRECT_HEADERS = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

print("Loading 10-sheet file...")
wb = load_workbook(file_path)

print(f"\n✅ Sheets ({len(wb.sheetnames)}): {wb.sheetnames}\n")

# Map old sheet names to new database table names
SHEET_MAP = {
    'Received': 'agency_received',
    'Returned at Agency': 'agency_returned',
    'Pending at Agency': 'pending_at_agency',
    'Forwarded to Bank': 'forwarded_to_bank',
    'Sanctioned': 'sanctioned_by_bank_no_of_proj',
    'MM Claimed': 'mm_claimed_no_of_proj',
    'MM Disbursement': 'mm_disbursement_no_of_proj',
    'Returned by Bank': 'returned_by_bank',
    'Pend MM Disbmt - Total': 'pending_at_bank_no_of_proj',
    'Pend MM Disbmt - Detail': 'pending_for_mm_disbursement'
}

print("Applying correct headers and database table names...\n")

for excel_name, db_name in SHEET_MAP.items():
    ws = wb[excel_name]
    
    # Write correct headers
    for col_idx, header in enumerate(CORRECT_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Rename sheet to database table name
    ws.title = db_name
    
    print(f"  ✓ {excel_name:35s} → {db_name}")

print(f"\nSaving...")
wb.save(file_path)
wb.close()

print(f"✅ Updated! File now has:")
print(f"   - 10 sheets with database table names")
print(f"   - All sheets with correct 68-column headers")
print(f"\nFile: {file_path}")
