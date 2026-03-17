import openpyxl
print('start')
wb=openpyxl.load_workbook('district_reports/KVIB_10_SHEETS_CLEANED.xlsx', read_only=False)
print('loaded', wb.sheetnames)
for s in wb.sheetnames:
    ws=wb[s]
    hdr=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    print(s, 'len hdr', len(hdr), 'max_col', ws.max_column, 'last 5', hdr[-7:])
print('done')
