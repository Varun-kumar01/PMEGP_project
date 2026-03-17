import openpyxl
import os

# Source file should be the final 10-sheet file you already built
SOURCE_FILE = "district_reports/KVIB_10_SHEETS_CLEANED.xlsx"
TARGET_FILE = "district_reports/FINAL_10_SHEETS_aligned.xlsx"

# DB columns from backend/pmegController.js
DB_COLUMNS = [
    "current_status", "under_process_agency_reason", "office_name", "agency_type", "state",
    "applicant_id", "applicant_name", "applicant_address", "applicant_mobile_no", "alternate_mobile_no",
    "email", "aadhar_no", "legal_status", "gender", "category", "special_category",
    "qualification", "date_of_birth", "age", "unit_location", "unit_address",
    "taluk_block", "unit_district", "industry_type", "product_desc_activity",
    "proposed_project_cost", "mm_involve", "financing_branch_ifsc_code", "financing_branch_address",
    "online_submission_date", "dltfec_meeting", "dltfec_meeting_place", "forwarding_date_to_bank",
    "bank_remarks", "date_of_documents_receiveda_at_bank",
    "project_cost_approved_ce", "project_cost_approved_wc", "project_cost_approved_total",
    "sanctioned_by_bank_date", "sanctioned_by_bank_ce", "sanctioned_by_bank_wc", "sanctioned_by_bank_total",
    "date_of_deposit_own_contribution", "own_contribution_amount_deposited",
    "covered_under_cgtsi", "date_of_loan_release", "loan_release_amount",
    "mm_claim_date", "mm_claim_amount", "remarks_for_mm_process_at_pmegp_co_mumbai",
    "mm_release_date", "mm_release_amount", "payment_status", "mm_disbursement_transaction_id", "fail_reason",
    "edp_training_center_name", "training_start_date", "training_end_date", "training_duration_days",
    "certificate_issue_date", "physical_verification_conducted_date", "physical_verification_status",
    "mm_final_adjustment_date", "mm_final_adjustment_amount",
    "tdr_account_no", "tdr_date", "year"
]

if not os.path.exists(SOURCE_FILE):
    raise FileNotFoundError(f"Source file not found: {SOURCE_FILE}")

wb = openpyxl.load_workbook(SOURCE_FILE)
print(f"Loaded workbook {SOURCE_FILE} with sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    first_row = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    print(f"Sheet '{sheet_name}' original cols={len(first_row)}")

    # overwrite first row with DB_COLUMNS
    for col_index, header in enumerate(DB_COLUMNS, 1):
        ws.cell(row=1, column=col_index, value=header)

    # clear header values beyond DB_COLUMNS length (in case workbook had extra columns)
    max_col = ws.max_column
    if max_col > len(DB_COLUMNS):
        for c in range(len(DB_COLUMNS) + 1, max_col + 1):
            ws.cell(row=1, column=c, value=None)

        # clear any columns beyond DB_COLUMNS length in existing data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=len(DB_COLUMNS) + 1, max_col=max_col):
            for cell in row:
                cell.value = None

# save
wb.save(TARGET_FILE)
print(f"Saved aligned workbook as {TARGET_FILE}")
