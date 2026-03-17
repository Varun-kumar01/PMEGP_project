import openpyxl
from itertools import islice
path='district_reports/FINAL_10_SHEETS_aligned.xlsx'
wb=openpyxl.load_workbook(path, read_only=True)
print('sheets',wb.sheetnames)
for s in wb.sheetnames:
    ws=wb[s]
    vals=[r[66] if len(r)>66 else None for r in islice(ws.iter_rows(min_row=2,max_row=5,values_only=True),3)]
    print(s, vals)
