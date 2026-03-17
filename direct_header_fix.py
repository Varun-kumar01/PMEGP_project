#!/usr/bin/env python3
"""Direct header fix - write headers directly to each sheet without deleting rows"""

from openpyxl import load_workbook

CORRECT = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

print("Loading workbook...")
wb = load_workbook(path)

for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f"[{sheet_idx:2d}] Writing headers to {sheet_name:30s} ... ", end="", flush=True)
    
    # Directly write headers to row 1, columns A-AH (1-68)
    for col_idx in range(1, 69):
        ws.cell(row=1, column=col_idx, value=CORRECT[col_idx-1])
    
    print("✓")

print(f"\nSaving workbook...")
wb.save(path)
wb.close()

print("✅ Done! Now verifying...")

# Quick verify
wb = load_workbook(path, read_only=True)
success_count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [str(ws.cell(1, c).value or '') for c in range(1, 69)]
    if all(h.lower() == e.lower() for h, e in zip(headers, CORRECT)):
        success_count += 1

wb.close()

print(f"✅ {success_count}/{len(wb.sheetnames)} sheets verified successfully!")
