#!/usr/bin/env python3
"""Quick check of Excel headers - read headers only, no data"""

from openpyxl import load_workbook
import os

# Exact 68 column names from database
EXPECTED = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year'
]

file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

print("Loading workbook (headers only, no data)...")
try:
    # Load with data_only=False to skip cached values, read_only=True for speed
    wb = load_workbook(file_path, data_only=False, read_only=True)
    print(f"✓ Loaded. Sheets: {wb.sheetnames}\n")
    
    ws = wb.active
    print(f"Checking first sheet: {ws.title}")
    
    # Read only row 1
    headers = []
    for col_idx in range(1, 70):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print(f"Found {len(headers)} headers\n")
    
    # Show first 5 and last 5
    print("First 5 headers:")
    for i, h in enumerate(headers[:5], 1):
        expected = EXPECTED[i-1] if i <= len(EXPECTED) else "N/A"
        match = "✓" if h.lower() == expected.lower() else "❌"
        print(f"  {i:2d}. {match} Got: '{h}' | Expect: '{expected}'")
    
    print("\nLast 5 headers:")
    start = len(headers) - 5
    for i, h in enumerate(headers[-5:], len(headers)-4):
        expected = EXPECTED[i-1] if i <= len(EXPECTED) else "N/A"
        match = "✓" if h.lower() == expected.lower() else "❌"
        print(f"  {i:2d}. {match} Got: '{h}' | Expect: '{expected}'")
    
    # Full comparison
    print(f"\nFull comparison:")
    if len(headers) != 68:
        print(f"❌ Column count mismatch: got {len(headers)}, expected 68")
    
    mismatches = []
    for i, (exp, got) in enumerate(zip(EXPECTED, headers), 1):
        if exp.lower() != got.lower():
            mismatches.append((i, exp, got))
    
    if mismatches:
        print(f"❌ {len(mismatches)} column mismatches found:")
        for col_num, exp, got in mismatches[:5]:
            print(f"  Col {col_num}: expected '{exp}' got '{got}'")
        if len(mismatches) > 5:
            print(f"  ... and {len(mismatches)-5} more")
    else:
        print(f"✅ ALL HEADERS MATCH!")
    
    wb.close()
    
except Exception as e:
    import traceback
    print(f"❌ ERROR: {e}")
    traceback.print_exc()
