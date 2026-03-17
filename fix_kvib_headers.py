import openpyxl
from openpyxl.utils import get_column_letter
import os
import shutil

# Database schema
db_columns = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type', 'state',
    'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no', 'alternate_mobile_no',
    'email', 'aadhar_no', 'legal_status', 'gender', 'category', 'special_category', 'qualification',
    'date_of_birth', 'age', 'unit_location', 'unit_address', 'taluk_block', 'unit_district',
    'industry_type', 'product_desc_activity', 'proposed_project_cost', 'mm_involve',
    'financing_branch_ifsc_code', 'financing_branch_address', 'online_submission_date', 'dltfec_meeting',
    'dltfec_meeting_place', 'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount', 'payment_status',
    'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name', 'training_start_date',
    'training_end_date', 'training_duration_days', 'certificate_issue_date', 'physical_verification_conducted_date',
    'physical_verification_status', 'mm_final_adjustment_date', 'mm_final_adjustment_amount',
    'tdr_account_no', 'tdr_date', 'year'
]

input_file = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023.xlsx"
output_file = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx"

try:
    print("=" * 80)
    print("FIXING EXCEL FILE HEADERS")
    print("=" * 80)
    print(f"\nInput file: {os.path.basename(input_file)}")
    
    # Load the workbook
    print("\n1. Loading workbook...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    print(f"   ✓ Loaded worksheet: {ws.title}")
    print(f"   ✓ Total rows: {ws.max_row}")
    print(f"   ✓ Total columns: {ws.max_column}")
    
    # Replace headers in row 1
    print(f"\n2. Replacing column headers with database schema ({len(db_columns)} columns)...")
    for col_idx, db_col_name in enumerate(db_columns, 1):
        ws.cell(row=1, column=col_idx).value = db_col_name
    
    # If there are more columns in the file than DB schema, clear them
    if ws.max_column > len(db_columns):
        print(f"   Clearing extra columns: {len(db_columns) + 1} to {ws.max_column}")
        for col_idx in range(len(db_columns) + 1, ws.max_column + 1):
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.value = None
    
    print(f"   ✓ Headers updated: {len(db_columns)} columns")
    
    # Save the file
    print(f"\n3. Saving fixed file...")
    wb.save(output_file)
    print(f"   ✓ File saved: {os.path.basename(output_file)}")
    
    # Verify the fix
    print(f"\n4. Verifying fix...")
    wb_check = openpyxl.load_workbook(output_file)
    ws_check = wb_check.active
    
    current_headers = []
    for col in range(1, 69):  # Check 68 columns
        cell = ws_check.cell(row=1, column=col)
        current_headers.append(str(cell.value).strip() if cell.value else "")
    
    # Count matches
    matches = sum(1 for i, header in enumerate(current_headers) if i < len(db_columns) and header == db_columns[i])
    
    print(f"   ✓ Verified columns: {matches}/{len(db_columns)} match")
    
    if matches == len(db_columns):
        print(f"\n{'='*80}")
        print(f"✓ SUCCESS! All {len(db_columns)} columns fixed correctly")
        print(f"{'='*80}")
        print(f"\nOutput file created:")
        print(f"  Location: {output_file}")
        print(f"  Size: {os.path.getsize(output_file) / (1024*1024):.2f} MB")
        print(f"\nYou can now use this file for upload.")
    else:
        print(f"\n⚠ Warning: Only {matches}/{len(db_columns)} columns match")
    
    wb_check.close()
    
except Exception as e:
    print(f"\n✗ Error: {e}")
    import traceback
    traceback.print_exc()
