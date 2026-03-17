#!/usr/bin/env python3
"""Definitively fix ALL 17 sheets with correct 68-column headers using openpyxl"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CORRECT_HEADERS = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year'
]

file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

print("Loading workbook...")
wb = load_workbook(file_path)
total_sheets = len(wb.sheetnames)

print(f"Loaded {total_sheets} sheets")
print(f"Sheet names: {wb.sheetnames}\n")

updated_count = 0
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f"[{idx:2d}/{total_sheets}] Updating {sheet_name:30s} ... ", end="", flush=True)
    
    # Delete row 1 if it exists
    try:
        ws.delete_rows(1, 1)
    except:
        pass  # Row might not exist
    
    # Insert new row at position 1
    ws.insert_rows(1, 1)
    
    # Write correct headers to columns A through AH (1-68)
    for col_idx, header_value in enumerate(CORRECT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header_value
    
    updated_count += 1
    print("✓ Done")

print(f"\nSaving workbook to: {file_path}")
print("(This may take a moment for the large file...)")
wb.save(file_path)
wb.close()

print(f"\n✅ SUCCESS: {updated_count} sheets updated with correct 68-column headers")
print("\nVerifying the fix...")

# Verify
wb = load_workbook(file_path, read_only=True)
all_correct = True
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [str(ws.cell(1, col).value or '').strip() for col in range(1, 69)]
    headers_match = all(h.lower() == e.lower() for h, e in zip(headers, CORRECT_HEADERS))
    if not headers_match:
        print(f"❌ {sheet_name} - Headers don't match")
        all_correct = False

if all_correct:
    print("✅ Verification passed: All sheets have correct headers!")
else:
    print("❌ Verification failed: Some sheets have incorrect headers")

wb.close()
