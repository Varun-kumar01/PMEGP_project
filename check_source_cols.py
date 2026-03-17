import openpyxl
wb=openpyxl.load_workbook('district_reports/KVIB_10_SHEETS_CLEANED.xlsx', read_only=True)
for s in wb.sheetnames:
    ws=wb[s]
    hdr=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    print(s, 'header count', len([x for x in hdr if x is not None]), 'max_col', ws.max_column, 'last header values', hdr[-3:])
