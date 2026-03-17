#!/usr/bin/env python3
import openpyxl

file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
wb = openpyxl.load_workbook(file_path)

print("=" * 100)
print("EXCEL FILE SHEETS ANALYSIS")
print("=" * 100)
print()
print(f"Total sheets: {len(wb.sheetnames)}")
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: '{sheet_name}'")
    print(f"  Rows: {ws.max_row}")
    print(f"  Columns: {ws.max_column}")
    print(f"  Row 1 (first 10 headers):")
    for col_idx in range(1, min(11, ws.max_column + 1)):
        cell = ws.cell(1, col_idx)
        print(f"    Col {col_idx}: {cell.value}")
    print()
