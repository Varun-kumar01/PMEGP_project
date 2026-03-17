#!/usr/bin/env python3
"""Delete unwanted sheets from FIXED file using openpyxl directly"""

from openpyxl import load_workbook
import shutil

source = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
backup = r'E:\kadhi\PMEGP_project\district_reports\KVIB_BACKUP_17sheets.xlsx'

# These are the 10 sheets to KEEP
SHEETS_TO_KEEP = {
    'Received',
    'Returned at Agency',
    'Pending at Agency',
    'Forwarded to Bank',
    'Sanctioned',
    'MM Claimed',
    'MM Disbursement',
    'Returned by Bank',
    'Pend MM Disbmt - Total',
    'Pend MM Disbmt - Detail'
}

print("Loading workbook...")
wb = load_workbook(source)
print(f"Current sheets ({len(wb.sheetnames)}): {wb.sheetnames}\n")

# Find sheets to DELETE
sheets_to_delete = [s for s in wb.sheetnames if s not in SHEETS_TO_KEEP]
print(f"Sheets to DELETE ({len(sheets_to_delete)}): {sheets_to_delete}\n")

# Create backup
print(f"Creating backup...")
shutil.copy(source, backup)
print(f"✓ Backed up to: {backup}\n")

# Delete unwanted sheets
print("Deleting unwanted sheets...")
for sheet_name in sheets_to_delete:
    try:
        sh = wb[sheet_name]
        wb.remove(sh)
        print(f"  ✓ Deleted: {sheet_name}")
    except Exception as e:
        print(f"  ❌ Failed to delete {sheet_name}: {e}")

print(f"\nRemaining sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

# Save
print(f"\nSaving...")
wb.save(source)
wb.close()

print(f"✅ Success! File now has {len(wb.sheetnames)} sheets")
print(f"✅ File: {source}")
