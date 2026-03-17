#!/usr/bin/env python3
"""Diagnose why Excel upload is failing"""
import openpyxl
import os

file_path = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_CORRECTED.xlsx"

if not os.path.exists(file_path):
    print(f"❌ File not found: {file_path}")
    exit(1)

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print("=" * 100)
    print("EXCEL FILE STRUCTURE ANALYSIS")
    print("=" * 100)
    print()
    
    # Check row 1
    print("ROW 1 (first row):")
    row1_cells = []
    for col_num in range(1, 10):
        cell = ws.cell(row=1, column=col_num)
        row1_cells.append(str(cell.value)[:40] if cell.value else "EMPTY")
    print(f"  Columns 1-9: {row1_cells}")
    print()
    
    # Check row 2
    print("ROW 2 (header row):")
    row2_cells = []
    for col_num in range(1, 10):
        cell = ws.cell(row=2, column=col_num)
        value = str(cell.value) if cell.value else "EMPTY"
        row2_cells.append(value[:40])
    print(f"  Columns 1-9: {row2_cells}")
    print()
    
    # Check if row 2 contains actual database column names
    expected_first_cols = ["id", "current_status", "under_process_agency_reason", "office_name", "agency_type"]
    row2_first_5 = []
    for col_num in range(1, 6):
        cell = ws.cell(row=2, column=col_num)
        value = str(cell.value).strip() if cell.value else ""
        row2_first_5.append(value)
    
    print("HEADER VALIDATION:")
    print(f"  Expected first 5 headers: {expected_first_cols}")
    print(f"  Actual first 5 headers:   {row2_first_5}")
    
    if row2_first_5 == expected_first_cols:
        print("  ✅ Headers match database schema!")
    else:
        print("  ❌ Headers DO NOT match!")
    print()
    
    # Check row 3 (first data row)
    print("ROW 3 (first data row):")
    row3_cells = []
    for col_num in range(1, 6):
        cell = ws.cell(row=3, column=col_num)
        value = str(cell.value)[:40] if cell.value else "EMPTY"
        row3_cells.append(value)
    print(f"  Columns 1-5: {row3_cells}")
    print()
    
    # Count total rows and columns
    max_row = ws.max_row
    max_col = ws.max_column
    
    print("FILE STATISTICS:")
    print(f"  Total rows (including header): {max_row}")
    print(f"  Total columns: {max_col}")
    print(f"  Data rows (excluding header): {max_row - 2}")
    print()
    
    if max_col >= 68:
        print("  ✅ File has at least 68 columns")
    else:
        print(f"  ❌ File only has {max_col} columns (need 68)")
    
    if max_row >= 3:
        print("  ✅ File has data rows")
    else:
        print("  ❌ File has no data rows")
    print()
    
    # Check consistency of columns across rows
    print("ROW CONSISTENCY CHECK:")
    row_col_distribution = {}
    for row_num in range(3, min(max_row + 1, 13)):  # Check first 10 data rows
        row = ws[row_num]
        col_count = 0
        for cell in row:
            if cell.value is not None:
                col_count = cell.column
        row_col_distribution[row_num] = col_count
    
    if row_col_distribution:
        print(f"  First 10 data rows column distribution:")
        for row_num, col_count in list(row_col_distribution.items())[:10]:
            print(f"    Row {row_num}: {col_count} columns")
    
    print()
    print("=" * 100)
    
except Exception as e:
    print(f"❌ Error reading file: {e}")
    import traceback
    traceback.print_exc()
