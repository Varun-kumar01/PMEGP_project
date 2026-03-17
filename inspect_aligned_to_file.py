import openpyxl
wb=openpyxl.load_workbook('district_reports/FINAL_10_SHEETS_aligned.xlsx', read_only=True)
with open('aligned_report.txt','w',encoding='utf-8') as out:
    out.write('sheets=' + str(wb.sheetnames) + '\n')
    for s in wb.sheetnames:
        ws=wb[s]
        hdr=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
        nonempty=[x for x in hdr if x is not None]
        out.write(f"{s}: len(hdr)={len(hdr)}, nonempty={len(nonempty)}, last5={hdr[-5:]}\n")
print('done')
