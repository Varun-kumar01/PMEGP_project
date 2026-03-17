import openpyxl
import os

# Check the existing FIXED file
fixed_file = r'district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

if os.path.exists(fixed_file):
    wb = openpyxl.load_workbook(fixed_file)
    ws = wb.active
    
    # Get all headers in row 1
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        headers.append(cell_val)
    
    print(f"File: {fixed_file}")
    print(f"Columns: {len(headers)}")
    print(f"Rows: {ws.max_row}")
    print(f"\nFirst 10 headers:")
    for i, h in enumerate(headers[:10], 1):
        print(f"  {i}: {h}")
    print(f"\nLast 5 headers:")
    for i, h in enumerate(headers[-5:], len(headers)-4):
        print(f"  {i}: {h}")
    
    # Check if "year" is in headers
    if "year" in headers:
        print(f"\n✓ 'year' column found at position {headers.index('year') + 1}")
    else:
        print(f"\n✗ 'year' column NOT found")
    
    # Check total count
    expected_count = 68
    if len(headers) == expected_count:
        print(f"✓ Correct: {len(headers)} columns (expected {expected_count})")
    else:
        print(f"⚠ Mismatch: {len(headers)} columns (expected {expected_count})")
else:
    print(f"File not found: {fixed_file}")
