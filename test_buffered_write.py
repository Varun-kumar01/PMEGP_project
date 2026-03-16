#!/usr/bin/env python3
"""Test buffered Excel write approach."""
import os
from io import BytesIO
import openpyxl

outdir = r"E:\kadhi\PMEGP_project\district_reports"
testfile = os.path.join(outdir, "TEST_BUFFERED_WRITE.xlsx")

print(f"Testing buffered write to: {testfile}")

try:
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestData"
    
    # Add data
    ws['A1'] = "Test Entry"
    ws['B1'] = 12345
    ws['C1'] = "Buffered write test"
    
    # Add more rows to test
    for i in range(2, 100):
        ws[f'A{i}'] = f"Row {i}"
        ws[f'B{i}'] = i * 100
    
    print(f"Before save: File exists = {os.path.exists(testfile)}")
    
    # Method 1: Direct save (old approach)
    # wb.save(testfile)
    
    # Method 2: Buffered write (new approach)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    with open(testfile, 'wb') as f:
        f.write(buffer.getvalue())
        f.flush()
        os.fsync(f.fileno())
    
    print(f"After save: File exists = {os.path.exists(testfile)}")
    
    if os.path.exists(testfile):
        size = os.path.getsize(testfile)
        print(f"SUCCESS: File created with {size} bytes")
        
        # Verify contents
        import pandas as pd
        df = pd.read_excel(testfile, sheet_name="TestData")
        print(f"Verification: DataFrame shape = {df.shape}")
        print(f"First row: {df.iloc[0].tolist()}")
    else:
        print("FAILURE: File was not created!")
        
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
