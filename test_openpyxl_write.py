#!/usr/bin/env python3
"""Minimal test to verify openpyxl can write files."""
import os
import openpyxl

outdir = r"E:\kadhi\PMEGP_project\district_reports"
testfile = os.path.join(outdir, "TEST_OPENPYXL.xlsx")

print(f"Output directory: {outdir}")
print(f"exists: {os.path.exists(outdir)}")
print(f"Output file path: {testfile}")

try:
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # Add data
    ws['A1'] = "Test Data"
    ws['B1'] = 123
    ws['C1'] = "openpyxl works" 
    
    print("Before save...")
    print(f"File exists before save: {os.path.exists(testfile)}")
    
    # Save
    wb.save(testfile)
    
    print("After save...")
    print(f"File exists after save: {os.path.exists(testfile)}")
    
    if os.path.exists(testfile):
        size = os.path.getsize(testfile)
        print(f"File size: {size} bytes")
        print("SUCCESS - File was created!")
    else:
        print("FAILURE - File does not exist after save!")
        
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
