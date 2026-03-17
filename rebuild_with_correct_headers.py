#!/usr/bin/env python3
"""Create new workbook with all sheets and correct headers, replacing original"""

from openpyxl import load_workbook
import shutil
import os

source = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
temp_new = r'E:\kadhi\PMEGP_project\district_reports\TEMP_NEW.xlsx'
backup = r'E:\kadhi\PMEGP_project\district_reports\BACKUP_OLD.xlsx'

CORRECT = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

print("Step 1: Reading source workbook...")
wb_source = load_workbook(source)
print(f"✓ Loaded {len(wb_source.sheetnames)} sheets")

print("\nStep 2: Creating new workbook with fixed headers...")
# Load source again to get raw data
wb_new = load_workbook(source, data_only=False)

for idx, sheet_name in enumerate(wb_new.sheetnames, 1):
    if idx <= 3 or idx > len(wb_new.sheetnames) - 2:  # Show progress on first 3 and last 2
        print(f"  [{idx:2d}] {sheet_name:30s} - ", end="", flush=True)
    
    ws = wb_new[sheet_name]
    
    # Write correct headers to row 1, columns A-AH
    for col_idx in range(1, 69):
        ws.cell(row=1, column=col_idx, value=CORRECT[col_idx-1])
    
    if idx <= 3 or idx > len(wb_new.sheetnames) - 2:
        print("✓")

print(f"  [{len(wb_new.sheetnames)//2:2d}] ... (processing {len(wb_new.sheetnames)} sheets)")

print("\nStep 3: Saving to temporary file...")
wb_new.save(temp_new)
print(f"✓ Saved to: {temp_new}")
wb_new.close()

print("\nStep 4: Replacing original file...")
if os.path.exists(backup):
    os.remove(backup)
shutil.move(source, backup)
print(f"✓ Original backed up to: {backup}")
shutil.move(temp_new, source)
print(f"✓ New file moved to: {source}")

print("\n" + "="*60)
print("Step 5: Final verification...")
wb_verify = load_workbook(source, read_only=True)
success_count = 0
total = len(wb_verify.sheetnames)

for sheet_name in wb_verify.sheetnames:
    ws = wb_verify[sheet_name]
    headers = [str(ws.cell(1, c).value or '') for c in range(1, 69)]
    if all(h.lower() == c.lower() for h, c in zip(headers, CORRECT)):
        success_count += 1

wb_verify.close()

print(f"\n✅ COMPLETE!")
print(f"✅ {success_count}/{total} sheets verified with correct headers")

if success_count == total:
    print("\n🎉 Excel file is ready for upload!")
else:
    print(f"\n⚠️ {total - success_count} sheets may still have issues")
