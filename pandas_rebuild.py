#!/usr/bin/env python3
"""Use pandas to directly rebuild Excel file with correct headers"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import shutil

source = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
temp_output = r'E:\kadhi\PMEGP_project\district_reports\TEMP_PANDAS.xlsx'

CORRECT_HEADERS = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

print(f"Excel file: {source}")
print(f"Opening with openpyxl to get sheet names...")

try:
    wb_src = load_workbook(source)
    sheet_names = wb_src.sheetnames
    wb_src.close()
    
    print(f"✓ Found {len(sheet_names)} sheets")
    print(f"  {sheet_names}\n")
    
    # Process with pandas and write directly
    with pd.ExcelWriter(temp_output, engine='openpyxl') as writer:
        for idx, sheet_name in enumerate(sheet_names, 1):
            print(f"[{idx:2d}] {sheet_name:30s} ... ", end="", flush=True)
            
            try:
                # Read sheet with no headers
                df = pd.read_excel(source, sheet_name=sheet_name, header=None, engine='openpyxl')
                
                # Update headers  (row 0 to be the correct ones)
                if len(df) > 0:
                    df.columns = CORRECT_HEADERS[:len(df.columns)]
                    # If fewer columns than expected, pad with empty
                    while len(df.columns) < 68:
                        df[len(df.columns)] = ""
                    # Or if more, truncate
                    df = df.iloc[:, :68]
                    df.columns = CORRECT_HEADERS
                
                # Write to temp with correct column names
                df.to_excel(writer, sheet_name=sheet_name, header=True, index=False)
                print("✓")
                
            except Exception as e:
                print(f"❌ Error: {str(e)[:40]}")
    
    print(f"\nSaved to: {temp_output}")
    
    # Verify temp file
    print("\nVerifying temp file...")
    wb_verify = load_workbook(temp_output, data_only=False)
    success = 0
    
    for sheet_name in wb_verify.sheetnames:
        ws = wb_verify[sheet_name]
        headers = [str(ws.cell(1, c).value or '') for c in range(1, 69)]
        if all(h.lower() == ch.lower() for h, ch in zip(headers, CORRECT_HEADERS)):
            success += 1
    
    wb_verify.close()
    
    if success == len(sheet_names):
        print(f"✅ Temp file verified! All {success} sheets have correct headers")
        
        # Replace original
        print("\nReplacing original file...")
        backup = source.replace('.xlsx', '_BACKUP_OLD.xlsx')
        if os.path.exists(backup):
            os.remove(backup)
        
        # Move source to backup
        shutil.move(source, backup)
        # Move temp to original location
        shutil.move(temp_output, source)
        print(f"✅ File replaced successfully!")
        print(f"\nFinal result: {success}/{len(sheet_names)} sheets ✓")
    else:
        print(f"⚠️ Only {success}/{len(sheet_names)} sheets verified")
        
except Exception as e:
    print(f"\n❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
