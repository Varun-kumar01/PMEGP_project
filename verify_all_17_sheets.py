#!/usr/bin/env python3
"""Verify ALL 17 sheets have correct 68-column headers"""

from openpyxl import load_workbook

EXPECTED = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

try:
    wb = load_workbook(file_path, data_only=False, read_only=True)
    
    all_good = True
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(ws.cell(1, col).value or '').strip() for col in range(1, 69)]
        
        # Check count and match
        count_ok = len([h for h in headers if h]) == 68
        headers_match = all(h.lower() == e.lower() for h, e in zip(headers, EXPECTED))
        
        status = "✅" if (count_ok and headers_match) else "❌"
        print(f"{status} {sheet_name:30s}")
        
        if not (count_ok and headers_match):
            all_good = False
    
    wb.close()
    
    print(f"\n{'='*50}")
    if all_good:
        print("✅ SUCCESS: All 17 sheets are correctly formatted!")
    else:
        print("❌ Some sheets still have issues")

except Exception as e:
    print(f"❌ ERROR: {e}")
