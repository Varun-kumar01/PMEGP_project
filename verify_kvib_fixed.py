import openpyxl

db_columns = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type', 'state',
    'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no', 'alternate_mobile_no',
    'email', 'aadhar_no', 'legal_status', 'gender', 'category', 'special_category', 'qualification',
    'date_of_birth', 'age', 'unit_location', 'unit_address', 'taluk_block', 'unit_district',
    'industry_type', 'product_desc_activity', 'proposed_project_cost', 'mm_involve',
    'financing_branch_ifsc_code', 'financing_branch_address', 'online_submission_date', 'dltfec_meeting',
    'dltfec_meeting_place', 'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount', 'payment_status',
    'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name', 'training_start_date',
    'training_end_date', 'training_duration_days', 'certificate_issue_date', 'physical_verification_conducted_date',
    'physical_verification_status', 'mm_final_adjustment_date', 'mm_final_adjustment_amount',
    'tdr_account_no', 'tdr_date', 'year'
]

fixed_file = r"district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx"

try:
    print("=" * 80)
    print("VERIFYING FIXED EXCEL FILE")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(fixed_file, read_only=True, data_only=True)
    ws = wb.active
    
    print(f"\nFile: {fixed_file}")
    print(f"Worksheet: {ws.title}")
    print(f"Rows: {ws.max_row:,}")
    print(f"Columns: {ws.max_column}")
    
    # Get headers
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        headers.append(str(cell.value).strip() if cell.value else "")
    
    print(f"\nExpected DB columns: {len(db_columns)}")
    print(f"Actual file columns: {len(headers)}\n")
    
    # Check first 10 and last 10
    print("FIRST 10 COLUMNS:")
    for i in range(min(10, len(headers))):
        match = "✓" if i < len(db_columns) and headers[i] == db_columns[i] else "✗"
        print(f"  {match} {i+1:2d}. Expected: '{db_columns[i]:40s}' | Actual: '{headers[i]}'")
    
    print("\nLAST 10 COLUMNS:")
    start_idx = max(len(headers) - 10, 0)
    for i in range(start_idx, len(headers)):
        expected = db_columns[i] if i < len(db_columns) else "N/A"
        match = "✓" if i < len(db_columns) and headers[i] == db_columns[i] else "✗"
        print(f"  {match} {i+1:2d}. Expected: '{expected:40s}' | Actual: '{headers[i]}'")
    
    # Overall check
    matches = sum(1 for i in range(min(len(headers), len(db_columns))) if headers[i] == db_columns[i])
    
    print(f"\n{'='*80}")
    print(f"SUMMARY:")
    print(f"  Columns matching DB schema: {matches}/{min(len(headers), len(db_columns))}")
    
    if matches == len(db_columns) and len(headers) >= len(db_columns):
        print(f"\n✓✓✓ PERFECT! All {len(db_columns)} columns are correctly named")
        print(f"File is ready for database upload!")
    elif len(headers) < len(db_columns):
        missing = len(db_columns) - len(headers)
        print(f"\n⚠ WARNING: Missing {missing} columns (have {len(headers)}, need {len(db_columns)})")
    else:
        print(f"\n⚠ Some column names don't match (matches: {matches}/{len(db_columns)})")
    
    wb.close()
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
