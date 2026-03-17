#!/usr/bin/env python3
"""Final clean save - load, verify headers, and save"""

from openpyxl import load_workbook
import time

path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

# Give any hanging processes time to release
time.sleep(3)

CORRECT = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

print("Loading workbook...")
try:
    wb = load_workbook(path)
    print(f"✓ Loaded {len(wb.sheetnames)} sheets")
    
    # Verify headers are already in memory from previous script
    for idx, sheet_name in enumerate(wb.sheetnames[:3], 1):  # Check first 3 as sample
        ws = wb[sheet_name]
        headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, 69)]
        is_correct = all(h.lower() == c.lower() for h, c in zip(headers, CORRECT))
        status = "✓" if is_correct else "❌"
        print(f"  {status} {sheet_name}")
    
    print("\nSaving workbook...")
    wb.save(path)
    print("✓ Saved successfully!")
    wb.close()
    
    # Quick verify after save
    print("\nVerifying saved file...")
    wb = load_workbook(path, read_only=True)
    success = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(ws.cell(1, c).value or '') for c in range(1, 69)]
        if all(h.lower() == c.lower() for h, c in zip(headers, CORRECT)):
            success += 1
    wb.close()
    
    print(f"✅ SUCCESS: {success}/{len(wb.sheetnames)} sheets verified!")
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
