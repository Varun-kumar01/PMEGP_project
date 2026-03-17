import openpyxl
wb=openpyxl.load_workbook('district_reports/FINAL_10_SHEETS_aligned.xlsx', read_only=True)
for s in wb.sheetnames:
    ws=wb[s]
    hdr=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    nonempty=[(i+1,v) for i,v in enumerate(hdr) if v is not None]
    print(s, 'len', len(hdr), 'nonempty', len(nonempty), nonempty[-6:])
