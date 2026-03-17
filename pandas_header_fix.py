#!/usr/bin/env python3
"""Fix headers using pandas - faster and more reliable for large files"""

import pandas as pd
from openpyxl import load_workbook

CORRECT = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

print("Opening workbook to get sheet names...")
wb = load_workbook(path, data_only=False)
sheet_names = wb.sheetnames
wb.close()

print(f"Processing {len(sheet_names)} sheets using pandas...\n")

for idx, sheet_name in enumerate(sheet_names, 1):
    try:
        print(f"[{idx:2d}] {sheet_name:30s} ... ", end="", flush=True)
        
        # Read with header=None to avoid pandas setting headers
        df = pd.read_excel(path, sheet_name=sheet_name, header=None, engine='openpyxl')
        
        # If rows exist, replace first row with correct headers
        if len(df) > 0:
            # Replace columns 0-67 with correct headers
            df.columns = list(range(len(df.columns)))
            # Get the first row (old headers)
            first_row = df.iloc[0, :].values if len(df) > 0 else None
            # Replace first row values with correct headers
            df.iloc[0, :68] = CORRECT
            
            # Write back to Excel
            with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
        
        print("✓")
    except Exception as e:
        print(f"❌ Error: {str(e)[:50]}")

print("\n✅ All sheets processed!")
