#!/usr/bin/env python3
"""Save the workbook properly after headers are already written"""

from openpyxl import load_workbook
import os
import time

path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
backup_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED_BACKUP.xlsx'

print("Attempting to lock and save the file...")
print(f"Target: {path}\n")

# Try to save with backoff
max_retries = 5
for attempt in range(max_retries):
    try:
        print(f"Attempt {attempt+1}/{max_retries}: ", end="", flush=True)
        
        # Load and immediately save - should be already modified
        wb = load_workbook(path)
        print(f"loaded ({len(wb.sheetnames)} sheets)... ", end="", flush=True)
        
        # Create backup first
        if os.path.exists(path) and not os.path.exists(backup_path):
            import shutil
            shutil.copy(path, backup_path)
            print(f"backed up... ", end="", flush=True)
        
        # Save
        wb.save(path)
        print("✓ SAVED!")
        wb.close()
        break
        
    except PermissionError as e:
        print(f"Permission denied (file in use)")
        if attempt < max_retries - 1:
            print(f"  Waiting 2 seconds before retry...")
            time.sleep(2)
        else:
            print(f"\n❌ Failed after {max_retries} attempts")
            import sys
            sys.exit(1)
    except Exception as e:
        print(f"❌ Error: {e}")
        import sys
        sys.exit(1)

print("\nNow verifying all sheets...")

# Verify
from openpyxl import load_workbook

CORRECT = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

wb = load_workbook(path, read_only=True)
all_good = True
correct_count = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [str(ws.cell(1, c).value or '') for c in range(1, 69)]
    is_match = all(h.lower() ==e.lower() for h, e in zip(headers, CORRECT))
    
    if is_match:
        print(f"✅ {sheet_name}")
        correct_count += 1
    else:
        print(f"❌ {sheet_name}")
        all_good = False

wb.close()

print(f"\n{'='*50}")
print(f"Result: {correct_count}/{len(wb.sheetnames)} sheets correct")
if all_good:
    print("✅ SUCCESS: All sheets are ready for upload!")
else:
    print("❌ Some sheets still have issues")
