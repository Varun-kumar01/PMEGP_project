#!/usr/bin/env python3
"""Verify that all sheets in FIXED.xlsx have correct 68 column headers"""

from openpyxl import load_workbook
import os

# Exact 68 column names from database schema
EXPECTED_COLUMNS = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code', 'financing_branch_address',
    'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place', 'forwarding_date_to_bank', 'bank_remarks',
    'date_of_documents_receiveda_at_bank', 'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total', 'sanctioned_by_bank_date',
    'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total', 'date_of_deposit_own_contribution', 'own_contribution_amount_deposited',
    'covered_under_cgtsi', 'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount', 'payment_status', 'mm_disbursement_transaction_id',
    'fail_reason', 'edp_training_center_name', 'training_start_date', 'training_end_date', 'training_duration_days',
    'certificate_issue_date', 'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date', 'mm_final_adjustment_amount',
    'tdr_account_no', 'tdr_date', 'year'
]

file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

print(f"Checking file: {file_path}\n")
print(f"File exists: {os.path.exists(file_path)}\n")

try:
    # Load workbook
    wb = load_workbook(file_path, data_only=False)
    print(f"✓ Workbook loaded successfully")
    print(f"✓ Total sheets: {len(wb.sheetnames)}")
    print(f"✓ Sheet names: {wb.sheetnames}\n")
    
    # Check each sheet
    all_correct = True
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*70}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*70}")
        
        ws = wb[sheet_name]
        
        # Get headers from row 1
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        print(f"Found {len(headers)} columns in row 1")
        
        # Check if headers match expected
        if len(headers) != 68:
            print(f"❌ ERROR: Expected 68 columns, found {len(headers)}")
            all_correct = False
        else:
            print(f"✓ Column count correct (68)")
        
        # Compare each header
        mismatches = []
        for i, (expected, actual) in enumerate(zip(EXPECTED_COLUMNS, headers), 1):
            if expected.lower() != actual.lower():
                mismatches.append((i, expected, actual))
        
        if mismatches:
            print(f"\n❌ Header mismatches found ({len(mismatches)}):")
            for col_num, expected, actual in mismatches[:10]:  # Show first 10
                print(f"  Col {col_num}: Expected '{expected}' but got '{actual}'")
            if len(mismatches) > 10:
                print(f"  ... and {len(mismatches) - 10} more mismatches")
            all_correct = False
        else:
            print(f"✓ All headers match database columns exactly")
        
        # Sample data rows
        data_row_count = ws.max_row - 1  # Exclude header
        if data_row_count > 0:
            print(f"✓ Data rows: {data_row_count}")
            
            # Check first data row for column count
            first_data_row = []
            for col_idx in range(1, 69):  # Check cols 1-68
                cell = ws.cell(row=2, column=col_idx)
                first_data_row.append(cell.value)
            
            non_empty = sum(1 for v in first_data_row if v is not None)
            print(f"  First data row: {non_empty} non-empty columns")
    
    print(f"\n{'='*70}")
    if all_correct:
        print("✅ ALL SHEETS VERIFIED - Headers are correct!")
    else:
        print("❌ Issues found - see above for details")
        print("\nWill create fix script to correct headers...")
    
    wb.close()

except Exception as e:
    print(f"❌ ERROR: {str(e)}")
    import traceback
    traceback.print_exc()
