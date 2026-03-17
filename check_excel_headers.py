import openpyxl
import os

# Path to the Excel file
excel_path = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023.xlsx"

# Database schema
db_columns = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type', 'state',
    'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no', 'alternate_mobile_no',
    'email', 'aadhar_no', 'legal_status', 'gender', 'category', 'special_category', 'qualification',
    'date_of_birth', 'age', 'unit_location', 'unit_address', 'taluk_block', 'unit_district',
    'industry_type', 'product_desc_activity', 'proposed_project_cost', 'mm_involve',
    'financing_branch_ifsc_code', 'financing_branch_address', 'online_submission_date', 'dltfec_meeting',
    'dltfec_meeting_place', 'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount', 'payment_status',
    'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name', 'training_start_date',
    'training_end_date', 'training_duration_days', 'certificate_issue_date', 'physical_verification_conducted_date',
    'physical_verification_status', 'mm_final_adjustment_date', 'mm_final_adjustment_amount',
    'tdr_account_no', 'tdr_date', 'year'
]

if os.path.exists(excel_path):
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        print(f"✓ File loaded: {os.path.basename(excel_path)}")
        print(f"  Worksheet: {ws.title}")
        print(f"  Total rows: {ws.max_row}")
        print(f"  Total columns: {ws.max_column}\n")
        
        # Get current headers from row 1
        current_headers = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                current_headers.append(str(cell.value).strip())
        
        print(f"Expected DB columns: {len(db_columns)}")
        print(f"Current Excel columns: {len(current_headers)}\n")
        
        print("=" * 80)
        print("COLUMN COMPARISON:")
        print("=" * 80)
        
        mismatches = []
        for i in range(max(len(db_columns), len(current_headers))):
            expected = db_columns[i] if i < len(db_columns) else "MISSING"
            actual = current_headers[i] if i < len(current_headers) else "MISSING"
            match = "✓" if expected == actual else "✗"
            
            if expected != actual:
                mismatches.append((i+1, expected, actual))
                print(f"{match} Col {i+1:2d}: Expected: '{expected:40s}' | Actual: '{actual}'")
            else:
                print(f"{match} Col {i+1:2d}: {expected}")
        
        print("\n" + "=" * 80)
        if mismatches:
            print(f"TOTAL MISMATCHES: {len(mismatches)}")
            for col_num, expected, actual in mismatches:
                print(f"  Column {col_num}: '{expected}' != '{actual}'")
        else:
            print("✓ ALL COLUMNS MATCH!")
        
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
else:
    print(f"✗ File not found: {excel_path}")
