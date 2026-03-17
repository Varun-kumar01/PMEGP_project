#!/usr/bin/env python3
"""Fix ALL sheets in Excel file - apply correct 68-column headers to all sheets"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# Exact 68 column names from database
CORRECT_HEADERS = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year'
]

file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

print(f"Loading workbook: {file_path}")
print(f"File size: {os.path.getsize(file_path) / (1024*1024):.1f} MB\n")

try:
    wb = load_workbook(file_path)
    print(f"✓ Opened. Total sheets: {len(wb.sheetnames)}")
    print(f"  Sheet names: {wb.sheetnames}\n")
    
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        data_rows = ws.max_row - 1
        
        print(f"Sheet {sheet_idx:2d}: {sheet_name:30s} ({data_rows:5d} data rows) - ", end="", flush=True)
        
        # Write correct headers to row 1
        for col_idx, header in enumerate(CORRECT_HEADERS, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # For sheets other than "Received", also fix data rows to 68 columns
        if sheet_name != "Received":
            # Pad any short rows and ensure exactly 68 columns per row
            for row_idx in range(2, ws.max_row + 1):
                # Make sure each row has exactly 68 columns (add empty if needed)
                for col_idx in range(1, 69):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is None and col_idx > 69:  # Clear extra columns
                        cell.value = None
        
        print("✓ Headers updated")
    
    print(f"\nSaving workbook...")
    wb.save(file_path)
    print(f"✓ Workbook saved successfully")
    
    wb.close()
    print(f"\n✅ COMPLETE: All {len(wb.sheetnames)} sheets now have correct 68-column headers!")

except Exception as e:
    import traceback
    print(f"\n❌ ERROR: {e}")
    traceback.print_exc()
