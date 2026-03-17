#!/usr/bin/env python3
"""Debug Excel file to see what backend will see"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

file_path = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_CORRECTED.xlsx"

if not os.path.exists(file_path):
    print(f"❌ File not found: {file_path}")
    exit(1)

def normalize_header(header):
    """Same function as backend"""
    if not header:
        return None
    return str(header).strip().lower().replace(" ", "_").replace("-", "_")

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print("=" * 120)
    print("EXCEL FILE DEBUG - What the backend will see")
    print("=" * 120)
    print()
    
    # Expected columns from backend
    expected_cols = [
        "current_status", "under_process_agency_reason", "office_name", "agency_type", "state",
        "applicant_id", "applicant_name", "applicant_address", "applicant_mobile_no", "alternate_mobile_no",
    ]
    
    print("ROW 1 HEADERS (first 15 columns):")
    print("-" * 120)
    
    row1_headers = []
    for col_num in range(1, 16):
        cell = ws.cell(row=1, column=col_num)
        raw_value = cell.value
        normalized = normalize_header(raw_value)
        row1_headers.append((col_num, raw_value, normalized))
        
        match_status = "✓ MATCH" if normalized in expected_cols else "✗ NO MATCH"
        print(f"  Col {col_num:2d}: Raw='{raw_value}' → Normalized='{normalized}' {match_status}")
    
    print()
    print("ROW 2 (first 10 values - should be data row 1):")
    print("-" * 120)
    for col_num in range(1, 11):
        cell = ws.cell(row=2, column=col_num)
        print(f"  Col {col_num:2d}: {cell.value}")
    
    print()
    print("ROW 3 (first 10 values - should be data row 2):")
    print("-" * 120)
    for col_num in range(1, 11):
        cell = ws.cell(row=3, column=col_num)
        print(f"  Col {col_num:2d}: {cell.value}")
    
    print()
    print("FILE STRUCTURE:")
    print("-" * 120)
    print(f"  Max row: {ws.max_row}")
    print(f"  Max column: {ws.max_column}")
    print()
    
    # Count how many headers match
    all_headers = []
    matches = 0
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        normalized = normalize_header(cell.value)
        all_headers.append(normalized)
        if normalized in expected_cols:
            matches += 1
    
    print(f"  Headers in Row 1 that match expected columns: {matches}/{len(expected_cols)}")
    
    # Check if row 2 has data
    row2_has_data = False
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_num)
        if cell.value is not None:
            row2_has_data = True
            break
    
    print(f"  Row 2 has data: {row2_has_data}")
    print()
    
    if matches == 0:
        print("⚠️  WARNING: No headers matched!")
        print()
        print("Row 1 headers (all):")
        for i, h in enumerate(all_headers[:20], 1):
            print(f"  {i}: {h}")
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
