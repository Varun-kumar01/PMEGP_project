import openpyxl
from openpyxl.utils import get_column_letter
import os
import tempfile
import shutil

# Database schema - exactly 68 columns
db_columns = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type', 'state',
    'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no', 'alternate_mobile_no',
    'email', 'aadhar_no', 'legal_status', 'gender', 'category', 'special_category', 'qualification',
    'date_of_birth', 'age', 'unit_location', 'unit_address', 'taluk_block', 'unit_district',
    'industry_type', 'product_desc_activity', 'proposed_project_cost', 'mm_involve',
    'financing_branch_ifsc_code', 'financing_branch_address', 'online_submission_date', 'dltfec_meeting',
    'dltfec_meeting_place', 'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount', 'payment_status',
    'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name', 'training_start_date',
    'training_end_date', 'training_duration_days', 'certificate_issue_date', 'physical_verification_conducted_date',
    'physical_verification_status', 'mm_final_adjustment_date', 'mm_final_adjustment_amount',
    'tdr_account_no', 'tdr_date', 'year'
]

input_file = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023.xlsx"
output_file = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx"

try:
    print("=" * 80)
    print("FIXING EXCEL FILE HEADERS (Optimized)")
    print("=" * 80)
    print(f"\nInput file: {os.path.basename(input_file)}")
    
    # Use read_only=True to avoid full parsing
    print("\n1. Loading workbook (read-only mode)...")
    wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active
    print(f"   ✓ Loaded worksheet: {ws.title}")
    print(f"   ✓ Max row: {ws.max_row}")
    print(f"   ✓ Max column: {ws.max_column}")
    
    # Get current headers
    current_headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            current_headers.append(str(cell.value).strip())
    
    print(f"   Current columns: {len(current_headers)}")
    wb.close()
    
    # Now open for writing in data_only=False mode to preserve structure
    print("\n2. Loading workbook for modification...")
    wb_write = openpyxl.load_workbook(input_file)
    ws_write = wb_write.active
    
    # Replace headers
    print(f"\n3. Replacing headers with database schema...")
    for col_idx, col_name in enumerate(db_columns, 1):
        ws_write.cell(row=1, column=col_idx).value = col_name
    
    # Unmerge any merged cells in the header row
    print(f"   Checking for merged cells...")
    merged_ranges = list(ws_write.merged_cells)
    for merged_range in merged_ranges:
        # Check if it intersects with row 1
        if merged_range.min_row == 1 or merged_range.max_row >= 1:
            try:
                ws_write.unmerge_cells(str(merged_range))
                print(f"    Unmerged: {merged_range}")
            except:
                pass
    
    # Save
    print(f"\n4. Saving fixed file...")
    wb_write.save(output_file)
    print(f"   ✓ File saved: {os.path.basename(output_file)}")
    
    size_mb = os.path.getsize(output_file) / (1024*1024)
    print(f"   Size: {size_mb:.2f} MB")
    
    # Verify
    print(f"\n5. Verifying fix...")
    wb_check = openpyxl.load_workbook(output_file, read_only=True, data_only=True)
    ws_check = wb_check.active
    
    verified_headers = []
    for col in range(1, min(69, ws_check.max_column + 1)):
        cell = ws_check.cell(row=1, column=col)
        verified_headers.append(str(cell.value).strip() if cell.value else "")
    
    matches = sum(1 for i, h in enumerate(verified_headers) if i < len(db_columns) and h == db_columns[i])
    
    print(f"   ✓ Columns verified: {matches}/{len(db_columns)} match DB schema")
    print(f"   ✓ Total rows: {ws_check.max_row:,}")
    
    if matches == len(db_columns) and ws_check.max_column >= len(db_columns):
        print(f"\n{'='*80}")
        print(f"✓ SUCCESS! Excel file fixed and ready for upload")
        print(f"{'='*80}")
        print(f"\nOutput: {output_file}")
    else:
        print(f"\n⚠ Partial fix - {matches} columns match")
    
    wb_check.close()
    wb_write.close()
    
except Exception as e:
    print(f"\n✗ Error: {e}")
    import traceback
    traceback.print_exc()
