import openpyxl, os
file='district_reports/FINAL_10_SHEETS_aligned.xlsx'
print('exists', os.path.exists(file))
if not os.path.exists(file):
    raise SystemExit(1)
wb=openpyxl.load_workbook(file, read_only=True)
print('sheets', wb.sheetnames)
for s in wb.sheetnames:
    ws=wb[s]
    hdr=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    nonempty=[x for x in hdr if x is not None]
    print(s, 'non-empty columns', len(nonempty), 'first 8', hdr[:8], 'last 5', hdr[-5:])
