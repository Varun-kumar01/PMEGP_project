#!/usr/bin/env python
"""
Fix the district detail Excel file with the 68 columns expected by the backend
"""
import openpyxl
import os
import shutil

# 68 columns required for agency detail data
DETAIL_COLS = [
    'id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type', 'state',
    'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no', 'alternate_mobile_no',
    'email', 'aadhar_no', 'legal_status', 'gender', 'category', 'special_category', 'qualification',
    'date_of_birth', 'age', 'unit_location', 'unit_address', 'taluk_block', 'unit_district',
    'industry_type', 'product_desc_activity', 'proposed_project_cost', 'mm_involve',
    'financing_branch_ifsc_code', 'financing_branch_address', 'online_submission_date', 'dltfec_meeting',
    'dltfec_meeting_place', 'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount', 'payment_status',
    'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name', 'training_start_date',
    'training_end_date', 'training_duration_days', 'certificate_issue_date', 'physical_verification_conducted_date',
    'physical_verification_status', 'mm_final_adjustment_date', 'mm_final_adjustment_amount',
    'tdr_account_no', 'tdr_date', 'year'
]

INPUT_FILE = r'district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023.xlsx'
OUTPUT_FILE = r'district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_CORRECT.xlsx'

def main():
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(f"File not found: {INPUT_FILE}")
    
    # Backup
    backup = INPUT_FILE.replace('.xlsx', '_BACKUP_LATEST.xlsx')
    if os.path.exists(backup):
        os.remove(backup)
    shutil.copy(INPUT_FILE, backup)
    
    print(f"Input: {INPUT_FILE}")
    print(f"Output: {OUTPUT_FILE}")
    
    # Load data
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    # Read all data (preserve all columns)
    data = []
    max_cols = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_cols):
        data.append([cell.value for cell in row])
    
    print(f"Original: {len(data)} rows, {max_cols} columns")
    
    # Create new workbook
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    
    # Set headers in row 1 (68 columns)
    for col_idx, header in enumerate(DETAIL_COLS, 1):
        ws_new.cell(row=1, column=col_idx, value=header)
    
    # Copy row 2 if exists (sub-headers)
    if len(data) > 1:
        for col_idx, val in enumerate(data[1][:68], 1):
            ws_new.cell(row=2, column=col_idx, value=val)
    
    # Copy data rows (skip first two rows)
    row_count = 0
    for row_idx, row_data in enumerate(data[2:], 3):
        for col_idx, val in enumerate(row_data[:68], 1):
            ws_new.cell(row=row_idx, column=col_idx, value=val)
        row_count += 1
    
    # Save
    wb_new.save(OUTPUT_FILE)
    
    print(f"✓ Created: {OUTPUT_FILE}")
    print(f"✓ Format: 68 columns (row 1) + data rows (row 3+)")
    print(f"✓ Data rows: {row_count}")
    print(f"✓ Backup: {backup}")
    print(f"\nStatus: READY FOR UPLOAD")

if __name__ == '__main__':
    main()
