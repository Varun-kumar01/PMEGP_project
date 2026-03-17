#!/usr/bin/env python3
"""Save the workbook with 10 sheets to a new file,  then replace original"""

from openpyxl import load_workbook
import shutil
import os
import time

source = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
temp_file = r'E:\kadhi\PMEGP_project\district_reports\TEMP_10SHEETS.xlsx'

SHEETS_TO_KEEP = {
    'Received', 'Returned at Agency', 'Pending at Agency', 'Forwarded to Bank',
    'Sanctioned', 'MM Claimed', 'MM Disbursement', 'Returned by Bank',
    'Pend MM Disbmt - Total', 'Pend MM Disbmt - Detail'
}

print("Loading workbook...")
wb = load_workbook(source)

# Delete unwanted sheets
sheets_to_delete = [s for s in wb.sheetnames if s not in SHEETS_TO_KEEP]
for sheet_name in sheets_to_delete:
    del wb[sheet_name]

print(f"✓ Removed 7 sheets, kept 10: {wb.sheetnames}\n")

print("Saving to temporary file...")
wb.save(temp_file)
wb.close()

print(f"✓ Saved to: {temp_file}")

# Replace original
print("\nReplacing original file...")
time.sleep(2)  # Wait for locks to clear
os.remove(source)
shutil.move(temp_file, source)

print(f"✅ Replaced! File now has 10 sheets")

# Verify
print("\nVerifying...")
wb_verify = load_workbook(source)
print(f"✅ Final sheets: {wb_verify.sheetnames}")
wb_verify.close()
