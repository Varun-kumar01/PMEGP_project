import openpyxl, os
SOURCE_FILE='district_reports/KVIB_10_SHEETS_CLEANED.xlsx'
DB_COLUMNS=[
    "current_status", "under_process_agency_reason", "office_name", "agency_type", "state",
    "applicant_id", "applicant_name", "applicant_address", "applicant_mobile_no", "alternate_mobile_no",
    "email", "aadhar_no", "legal_status", "gender", "category", "special_category",
    "qualification", "date_of_birth", "age", "unit_location", "unit_address",
    "taluk_block", "unit_district", "industry_type", "product_desc_activity",
    "proposed_project_cost", "mm_involve", "financing_branch_ifsc_code", "financing_branch_address",
    "online_submission_date", "dltfec_meeting", "dltfec_meeting_place", "forwarding_date_to_bank",
    "bank_remarks", "date_of_documents_receiveda_at_bank",
    "project_cost_approved_ce", "project_cost_approved_wc", "project_cost_approved_total",
    "sanctioned_by_bank_date", "sanctioned_by_bank_ce", "sanctioned_by_bank_wc", "sanctioned_by_bank_total",
    "date_of_deposit_own_contribution", "own_contribution_amount_deposited",
    "covered_under_cgtsi", "date_of_loan_release", "loan_release_amount",
    "mm_claim_date", "mm_claim_amount", "remarks_for_mm_process_at_pmegp_co_mumbai",
    "mm_release_date", "mm_release_amount", "payment_status", "mm_disbursement_transaction_id", "fail_reason",
    "edp_training_center_name", "training_start_date", "training_end_date", "training_duration_days",
    "certificate_issue_date", "physical_verification_conducted_date", "physical_verification_status",
    "mm_final_adjustment_date", "mm_final_adjustment_amount",
    "tdr_account_no", "tdr_date", "year"
]
wb=openpyxl.load_workbook(SOURCE_FILE)
for sheet_name in wb.sheetnames:
    ws=wb[sheet_name]
    first_row=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    print('before',sheet_name,'len',len(first_row),'non-empty',len([x for x in first_row if x is not None]),'last7',first_row[-7:])
    max_col=ws.max_column
    for i,h in enumerate(DB_COLUMNS,1):
        ws.cell(row=1,column=i,value=h)
    if max_col>len(DB_COLUMNS):
        for c in range(len(DB_COLUMNS)+1,max_col+1):
            ws.cell(row=1,column=c,value=None)
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=len(DB_COLUMNS)+1,max_col=max_col):
            for cell in row:
                cell.value=None
    first_row2=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    print('after',sheet_name,'len',len(first_row2),'non-empty',len([x for x in first_row2 if x is not None]),'last7',first_row2[-7:])
