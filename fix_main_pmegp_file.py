import openpyxl
from openpyxl.utils import get_column_letter
import os
import shutil

# The 18 columns expected by the backend uploadPmegData function
required_columns = [
    'rowNo',              # Column 1
    'name',               # Column 2  
    'agencyReceived',     # Column 3
    'agencyReturned',     # Column 4
    'Pending_At_Agency',  # Column 5
    'Forwarded_to_Bank',  # Column 6
    'sanctionedPrj',      # Column 7
    'sanctionedLakh',     # Column 8
    'claimedPrj',         # Column 9
    'claimedLakh',        # Column 10
    'disbursementPrj',    # Column 11
    'disbursementLakh',   # Column 12
    'bankReturned',       # Column 13
    'pendingBankPrj',     # Column 14
    'pendingBankLakh',    # Column 15
    'pendingDisbursementPrj',  # Column 16
    'pendingDisbursementLakh',  # Column 17
    'year'                # Column 18
]

input_file = r"excel/pmegp_TG_KVIB_district_pipeline_01APR2022_to_31MAR2023.xlsx"
output_file = r"excel/pmegp_TG_KVIB_district_pipeline_01APR2022_to_31MAR2023_FIXED.xlsx"

try:
    print("=" * 80)
    print("FIXING MAIN PMEGP FILE - SETTING CORRECT COLUMNS")
    print("=" * 80)
    print(f"\nInput file: {os.path.basename(input_file)}")
    
    # Create backup
    backup_file = input_file.replace('.xlsx', '_BACKUP.xlsx')
    if os.path.exists(backup_file):
        os.remove(backup_file)
    shutil.copy(input_file, backup_file)
    print(f"✓ Backup created: {os.path.basename(backup_file)}")
    
    # Load the workbook
    print("\n1. Loading workbook...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    print(f"   Sheet: {ws.title}")
    print(f"   Total rows: {ws.max_row}")
    print(f"   Total columns: {ws.max_column}")
    
    # Read and preserve all data (columns 1-20 in case there are extra columns)
    print("\n2. Preserving existing data...")
    all_data = []
    for row_idx in range(1, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(cell.value)
        all_data.append(row_data)
    print(f"   ✓ Preserved {len(all_data)} rows")
    
    # Create fresh workbook with correct structure
    print("\n3. Creating fixed workbook with 18 columns...")
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    
    # Row 1: Main Headers
    print("   Setting Row 1 headers:")
    for col_idx, header in enumerate(required_columns, 1):
        ws_new.cell(row=1, column=col_idx).value = header
        print(f"     Col {col_idx}: {header}")
    
    # Row 2: Sub-headers (keep as-is or make blank)
    print("   Setting Row 2 (keeping original if exists)...")
    if len(all_data) > 1:
        for col_idx in range(1, min(19, len(all_data[1]) + 1)):
            ws_new.cell(row=2, column=col_idx).value = all_data[1][col_idx - 1]
    
    # Rows 3+: Data rows
    print("   Copying data rows (starting from row 3)...")
    data_start = 3
    copied_rows = 0
    
    if len(all_data) > 2:
        for src_row_idx in range(2, len(all_data)):  # Skip header rows
            for col_idx in range(1, min(19, len(all_data[src_row_idx]) + 1)):
                src_value = all_data[src_row_idx][col_idx - 1]
                ws_new.cell(row=data_start, column=col_idx).value = src_value
            data_start += 1
            copied_rows += 1
    
    print(f"   ✓ Copied {copied_rows} data rows")
    
    # Save fixed file
    print(f"\n4. Saving fixed file...")
    wb_new.save(output_file)
    print(f"   ✓ Saved: {os.path.basename(output_file)}")
    
    # Verify
    print(f"\n5. Verifying fix...")
    wb_check = openpyxl.load_workbook(output_file)
    ws_check = wb_check.active
    
    # Check headers
    headers_ok = True
    for col_idx, expected_header in enumerate(required_columns, 1):
        actual = ws_check.cell(row=1, column=col_idx).value
        if actual != expected_header:
            headers_ok = False
            print(f"   ✗ Col {col_idx}: Expected '{expected_header}', got '{actual}'")
    
    if headers_ok:
        print(f"   ✓ All 18 headers are correct")
        print(f"   ✓ Total data rows: {ws_check.max_row - 2}")  # minus 2 header rows
    
    wb_check.close()
    
    # Final summary
    print(f"\n{'='*80}")
    print(f"✓✓✓ SUCCESS! PMEGP file fixed!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_file}")
    print(f"Structure: Headers (rows 1-2) + Data (rows 3+)")
    print(f"Columns: 18 (matching backend requirements)")
    print(f"\nThe file is now ready for upload. The backend should:")
    print(f"1. Read column A as 'rowNo'")
    print(f"2. Read column B as 'name' (used to identify non-empty rows)")
    print(f"3. Read columns C-R for other data")
    print(f"4. Skip any null 'name' values")
    
except Exception as e:
    print(f"\n✗ Error: {e}")
    import traceback
    traceback.print_exc()
