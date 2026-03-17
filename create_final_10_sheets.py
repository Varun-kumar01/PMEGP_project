#!/usr/bin/env python3
"""Create new Excel file with only 10 required sheets and correct 68-column headers"""

import pandas as pd
from openpyxl import load_workbook
import shutil
import os

source = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
output = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FINAL.xlsx'

# These are the 10 sheets to keep (database table names)
REQUIRED_SHEETS = [
    'agency_received',
    'agency_returned', 
    'pending_at_agency',
    'forwarded_to_bank',
    'sanctioned_by_bank_no_of_proj',
    'mm_claimed_no_of_proj',
    'mm_disbursement_no_of_proj',
    'returned_by_bank',
    'pending_at_bank_no_of_proj',
    'pending_for_mm_disbursement'
]

# Mapping from Excel sheet names (as they currently exist) to database table names
SHEET_NAME_MAPPING = {
    'Received': 'agency_received',
    'Returned at Agency': 'agency_returned',
    'Pending at Agency': 'pending_at_agency',
    'Forwarded to Bank': 'forwarded_to_bank',
    'Sanctioned': 'sanctioned_by_bank_no_of_proj',
    'MM Claimed': 'mm_claimed_no_of_proj',
    'MM Disbursement': 'mm_disbursement_no_of_proj',
    'Returned by Bank': 'returned_by_bank',
    'Pending at Bank no of proj': 'pending_at_bank_no_of_proj',  # Try variations
    'Pend MM Disbmt - Total': 'pending_at_bank_no_of_proj',      # Fallback
    'Pend MM Disbmt - Detail': 'pending_for_mm_disbursement'
}

CORRECT_HEADERS = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
    'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
    'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
    'category', 'special_category', 'qualification', 'date_of_birth', 'age',
    'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
    'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
    'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
    'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
    'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
    'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
    'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
    'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
    'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
    'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
    'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
    'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
    'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']

print("Step 1: Loading current file to get sheet names...")
try:
    wb_src = load_workbook(source)
    current_sheets = wb_src.sheetnames
    wb_src.close()
    print(f"✓ Current sheets ({len(current_sheets)}): {current_sheets}\n")
    
    # Identify which current sheets correspond to our required tables
    print("Step 2: Mapping sheets...")
    sheets_to_process = {}
    
    for current_name, db_name in SHEET_NAME_MAPPING.items():
        if current_name in current_sheets and db_name in REQUIRED_SHEETS:
            sheets_to_process[current_name] = db_name
            print(f"  ✓ '{current_name}' → '{db_name}'")
    
    print(f"\nFound {len(sheets_to_process)} / 10 required sheets")
    
    # If we're missing some, list missing ones
    mapped_db_names = set(sheets_to_process.values())
    missing = set(REQUIRED_SHEETS) - mapped_db_names
    if missing:
        print(f"\n⚠️ Missing sheets (will need manual assignment):")
        for m in missing:
            print(f"  - {m}")
        print(f"\nAvailable unmapped sheets:")
        for s in current_sheets:
            if s not in sheets_to_process:
                print(f"  - {s}")
    
    print(f"\nStep 3: Creating new Excel file with {len(sheets_to_process)} sheets...")
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for idx, (excel_sheet_name, db_sheet_name) in enumerate(sheets_to_process.items(), 1):
            print(f"  [{idx:2d}] {db_sheet_name:35s} ... ", end="", flush=True)
            
            try:
                # Read from source
                df = pd.read_excel(source, sheet_name=excel_sheet_name, header=None, engine='openpyxl')
                
                # Set correct columns
                if len(df) > 0:
                    # Ensure 68 columns
                    while len(df.columns) < 68:
                        df[len(df.columns)] = ""
                    df = df.iloc[:, :68]
                    df.columns = CORRECT_HEADERS
                
                # Write to new file with database table name as sheet name
                df.to_excel(writer, sheet_name=db_sheet_name, header=True, index=False)
                print("✓")
                
            except Exception as e:
                print(f"❌ {str(e)[:30]}")
    
    print(f"\n✓ New file created: {output}")
    
    # Verify
    print("\nStep 4: Verifying...")
    wb_verify = load_workbook(output)
    
    print(f"✓ Sheets in new file: {wb_verify.sheetnames}")
    
    success = 0
    for sheet_name in wb_verify.sheetnames:
        ws = wb_verify[sheet_name]
        headers = [str(ws.cell(1, c).value or '') for c in range(1, 69)]
        if all(h.lower() == ch.lower() for h, ch in zip(headers, CORRECT_HEADERS)):
            success += 1
            print(f"  ✅ {sheet_name}")
        else:
            print(f"  ❌ {sheet_name} - headers don't match")
    
    wb_verify.close()
    
    print(f"\n{'='*60}")
    if success == len(sheets_to_process):
        print(f"✅ SUCCESS: {success} sheets with correct headers!")
        print(f"\n📁 New file: {output}")
        print(f"\nYou can now upload this file with the 10 required tables.")
    else:
        print(f"⚠️ Only {success}/{len(sheets_to_process)} sheets verified")
    
except Exception as e:
    print(f"\n❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
