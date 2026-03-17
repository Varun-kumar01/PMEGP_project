from openpyxl import load_workbook
path='E:\\kadhi\\PMEGP_project\\district_reports\\FINAL_10_SHEETS.xlsx'
wb=load_workbook(path,read_only=True)
print('sheets',wb.sheetnames)
for s in wb.sheetnames:
    ws=wb[s]
    headers=[(i, ws.cell(1,i).value) for i in range(1,90) if ws.cell(1,i).value is not None]
    print(s, len(headers), headers[:5], '...', headers[-5:])
wb.close()
