import openpyxl

try:
    wb = openpyxl.load_workbook('excel/pmegp_TG_KVIB_district_pipeline_01APR2022_to_31MAR2023.xlsx')
    ws = wb.active
    
    h1 = [ws.cell(1,i).value for i in range(1, 20)]
    h2 = [ws.cell(2,i).value for i in range(1, 20)]
    h3 = [ws.cell(3,i).value for i in range(1, 20)]
    
    with open('excel_structure.txt', 'w') as f:
        f.write("=== ROW 1 ===\n")
        for i, val in enumerate(h1, 1):
            f.write(f"Col {i}: {val}\n")
        
        f.write("\n=== ROW 2 ===\n")
        for i, val in enumerate(h2, 1):
            f.write(f"Col {i}: {val}\n")
        
        f.write("\n=== ROW 3 (first data) ===\n")
        for i, val in enumerate(h3, 1):
            f.write(f"Col {i}: {val}\n")
            
        f.write(f"\nTotal rows: {ws.max_row}\n")
        f.write(f"Total columns: {ws.max_column}\n")
    
    print("Done! Check excel_structure.txt")
except Exception as e:
    with open('excel_structure.txt', 'w') as f:
        f.write(f"Error: {e}\n")
        import traceback
        f.write(traceback.format_exc())
