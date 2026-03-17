import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('excel/pmegp_TG_KVIB_district_pipeline_01APR2022_to_31MAR2023.xlsx')
    ws = wb.active
    
    print("=== EXCEL FILE STRUCTURE ===")
    print(f"Sheet Name: {wb.sheetnames[0]}")
    print(f"Total rows: {ws.max_row}")
    print(f"Total columns: {ws.max_column}")
    
    print("\n=== ROW 1 (Header 1) ===")
    row1 = []
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        row1.append(cell_val)
        print(f"Col {col}: {cell_val}")
    
    print("\n=== ROW 2 (Header 2) ===")
    row2 = []
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=2, column=col).value
        row2.append(cell_val)
        print(f"Col {col}: {cell_val}")
    
    print("\n=== ROW 3 (First Data Row) ===")
    row3 = []
    for col in range(1, min(ws.max_column + 1, 20)):
        cell_val = ws.cell(row=3, column=col).value
        row3.append(cell_val)
        print(f"Col {col}: {cell_val}")
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
