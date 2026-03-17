#!/usr/bin/env python
"""
Fix the main PMEGP district pipeline Excel file.
Updates it with the 18 columns expected by the backend.
"""
import openpyxl
import os
import shutil

# Expected columns for main data file
REQUIRED_COLS = [
    'rowNo', 'name', 'agencyReceived', 'agencyReturned', 'Pending_At_Agency',
    'Forwarded_to_Bank', 'sanctionedPrj', 'sanctionedLakh', 'claimedPrj',
    'claimedLakh', 'disbursementPrj', 'disbursementLakh', 'bankReturned',
    'pendingBankPrj', 'pendingBankLakh', 'pendingDisbursementPrj', 'pendingDisbursementLakh', 'year'
]

INPUT_FILE = 'excel/pmegp_TG_KVIB_district_pipeline_01APR2022_to_31MAR2023.xlsx'
OUTPUT_FILE = 'excel/pmegp_TG_KVIB_district_pipeline_01APR2022_to_31MAR2023_FIXED.xlsx'

def main():
    # Check input file
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(f"File not found: {INPUT_FILE}")
    
    # Backup
    backup = INPUT_FILE.replace('.xlsx', '_BACKUP.xlsx')
    shutil.copy(INPUT_FILE, backup)
    
    # Load data
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    # Read all data
    data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=min(ws.max_column, 20)):
        data.append([cell.value for cell in row])
    
    # Create new workbook
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    
    # Set headers in row 1
    for col_idx, header in enumerate(REQUIRED_COLS, 1):
        ws_new.cell(row=1, column=col_idx, value=header)
    
    # Copy row 2 if exists (sub-headers)
    if len(data) > 1:
        for col_idx, val in enumerate(data[1][:18], 1):
            ws_new.cell(row=2, column=col_idx, value=val)
    
    # Copy data rows (skip first two rows)
    for row_idx, row_data in enumerate(data[2:], 3):
        for col_idx, val in enumerate(row_data[:18], 1):
            ws_new.cell(row=row_idx, column=col_idx, value=val)
    
    # Save
    wb_new.save(OUTPUT_FILE)
    
    # Result
    with open('pmegp_fix_result.txt', 'w') as f:
        f.write(f"SUCCESS\n")
        f.write(f"Input: {INPUT_FILE}\n")
        f.write(f"Output: {OUTPUT_FILE}\n")
        f.write(f"Backup: {backup}\n")
        f.write(f"Total data rows: {len(data) - 2}\n")
        f.write(f"Columns: 18\n")

if __name__ == '__main__':
    main()
