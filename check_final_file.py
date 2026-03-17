#!/usr/bin/env python3
"""Check what's in FINAL_10_SHEETS.xlsx"""

from openpyxl import load_workbook

file_path = r'E:\kadhi\PMEGP_project\district_reports\FINAL_10_SHEETS.xlsx'

try:
    wb = load_workbook(file_path)
    print(f"File: FINAL_10_SHEETS.xlsx")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_count = ws.max_row - 1  # Subtract header
        col_count = ws.max_column
        headers = []
        for col in range(1, min(10, col_count+1)):
            headers.append(str(ws.cell(1, col).value or ''))
        
        print(f"✓ {sheet_name:35s} - {row_count:6d} rows, {col_count:2d} cols")
        print(f"  Headers: {', '.join(headers[:3])}...")
    
    wb.close()
    
except Exception as e:
    print(f"❌ Error: {e}")
