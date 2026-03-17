#!/usr/bin/env python3
import openpyxl
import sys

try:
    file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
    wb = openpyxl.load_workbook(file_path)
    
    output = []
    output.append("=" * 100)
    output.append("EXCEL FILE SHEETS ANALYSIS")
    output.append("=" * 100)
    output.append("")
    output.append(f"Total sheets: {len(wb.sheetnames)}")
    output.append(f"Sheet names: {wb.sheetnames}")
    output.append("")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.append(f"Sheet: '{sheet_name}'")
        output.append(f"  Rows: {ws.max_row}")
        output.append(f"  Columns: {ws.max_column}")
        output.append(f"  Row 1 (first 15 headers):")
        for col_idx in range(1, min(16, ws.max_column + 1)):
            cell = ws.cell(1, col_idx)
            output.append(f"    {col_idx}: {cell.value}")
        output.append("")
    
    # Write to file
    with open(r'E:\kadhi\PMEGP_project\sheets_analysis.txt', 'w') as f:
        f.write('\n'.join(output))
    
    print("Analysis written to sheets_analysis.txt")
    
except Exception as e:
    print(f"ERROR: {e}")
    sys.exit(1)
