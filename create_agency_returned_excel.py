#!/usr/bin/env python3
"""Create Excel file with correct column headers"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

# The exact 68 columns the database expects (excluding auto-increment id)
correct_columns = [
    "id", "current_status", "under_process_agency_reason", "office_name", "agency_type", "state",
    "applicant_id", "applicant_name", "applicant_address", "applicant_mobile_no", "alternate_mobile_no",
    "email", "aadhar_no", "legal_status", "gender", "category", "special_category",
    "qualification", "date_of_birth", "age", "unit_location", "unit_address",
    "taluk_block", "unit_district", "industry_type", "product_desc_activity",
    "proposed_project_cost", "mm_involve", "financing_branch_ifsc_code", "financing_branch_address",
    "online_submission_date", "dltfec_meeting", "dltfec_meeting_place", "forwarding_date_to_bank",
    "bank_remarks", "date_of_documents_receiveda_at_bank",
    "project_cost_approved_ce", "project_cost_approved_wc", "project_cost_approved_total",
    "sanctioned_by_bank_date", "sanctioned_by_bank_ce", "sanctioned_by_bank_wc", "sanctioned_by_bank_total",
    "date_of_deposit_own_contribution", "own_contribution_amount_deposited",
    "covered_under_cgtsi", "date_of_loan_release", "loan_release_amount",
    "mm_claim_date", "mm_claim_amount", "remarks_for_mm_process_at_pmegp_co_mumbai",
    "mm_release_date", "mm_release_amount", "payment_status", "mm_disbursement_transaction_id", "fail_reason",
    "edp_training_center_name", "training_start_date", "training_end_date", "training_duration_days",
    "certificate_issue_date", "physical_verification_conducted_date", "physical_verification_status",
    "mm_final_adjustment_date", "mm_final_adjustment_amount",
    "tdr_account_no", "tdr_date", "year"
]

input_file = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx"
output_file = r"E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_AGENCY_RETURNED_READY.xlsx"

print("Processing Excel file...")
print(f"Input: {input_file}")
print(f"Output: {output_file}")
print()

try:
    # Load the source workbook
    print("Loading source file...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    print(f"Source file has {ws.max_row} rows and {ws.max_column} columns")
    
    # Create new workbook with correct headers
    print("Creating new workbook with correct headers...")
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # Write correct headers in row 1
    print(f"Writing {len(correct_columns)} headers to row 1...")
    for col_idx, col_name in enumerate(correct_columns, 1):
        cell = new_ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Copy data rows from source (skip header row)
    print(f"Copying data rows...")
    data_row_count = 0
    for src_row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(correct_columns) + 1):
            src_cell = ws.cell(row=src_row_idx, column=col_idx)
            dst_cell = new_ws.cell(row=src_row_idx, column=col_idx)
            dst_cell.value = src_cell.value
        data_row_count += 1
    
    # Adjust column widths
    print("Adjusting column widths...")
    for col_idx in range(1, len(correct_columns) + 1):
        new_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
    
    # Save new workbook
    print(f"Saving file...")
    new_wb.save(output_file)
    
    print()
    print("=" * 100)
    print(f"✅ SUCCESS!")
    print(f"Output file created: {output_file}")
    print(f"Headers: {len(correct_columns)} columns")
    print(f"Data rows: {data_row_count}")
    print(f"Total rows: {data_row_count + 1}")
    print("=" * 100)
    
except Exception as e:
    print(f"❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
