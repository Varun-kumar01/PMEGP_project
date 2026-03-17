#!/usr/bin/env python3
"""Save workbook using copy/rename approach to avoid locks"""

from openpyxl import load_workbook
import shutil
import os

path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'
temp_path = r'E:\kadhi\PMEGP_project\district_reports\TEMP_FIXED.xlsx'

print("Creating temporary copy and modifying...")

try:
    # Load workbook (already has headers in memory from previous run)
    wb = load_workbook(path)
    print(f"✓ Loaded workbook with {len(wb.sheetnames)} sheets")
    
    # Save to temp file
    wb.save(temp_path)
    print(f"✓ Saved to temporary file: {temp_path}")
    wb.close()
    
    # Replace original with temp
    os.remove(path)
    shutil.move(temp_path, path)
    print(f"✓ Replaced original file")
    
    print("\nVerifying...")
    
    # Verify final result
    CORRECT = ['id', 'current_status', 'under_process_agency_reason', 'office_name', 'agency_type',
        'state', 'applicant_id', 'applicant_name', 'applicant_address', 'applicant_mobile_no',
        'alternate_mobile_no', 'email', 'aadhar_no', 'legal_status', 'gender',
        'category', 'special_category', 'qualification', 'date_of_birth', 'age',
        'unit_location', 'unit_address', 'taluk_block', 'unit_district', 'industry_type',
        'product_desc_activity', 'proposed_project_cost', 'mm_involve', 'financing_branch_ifsc_code',
        'financing_branch_address', 'online_submission_date', 'dltfec_meeting', 'dltfec_meeting_place',
        'forwarding_date_to_bank', 'bank_remarks', 'date_of_documents_receiveda_at_bank',
        'project_cost_approved_ce', 'project_cost_approved_wc', 'project_cost_approved_total',
        'sanctioned_by_bank_date', 'sanctioned_by_bank_ce', 'sanctioned_by_bank_wc', 'sanctioned_by_bank_total',
        'date_of_deposit_own_contribution', 'own_contribution_amount_deposited', 'covered_under_cgtsi',
        'date_of_loan_release', 'loan_release_amount', 'mm_claim_date', 'mm_claim_amount',
        'remarks_for_mm_process_at_pmegp_co_mumbai', 'mm_release_date', 'mm_release_amount',
        'payment_status', 'mm_disbursement_transaction_id', 'fail_reason', 'edp_training_center_name',
        'training_start_date', 'training_end_date', 'training_duration_days', 'certificate_issue_date',
        'physical_verification_conducted_date', 'physical_verification_status', 'mm_final_adjustment_date',
        'mm_final_adjustment_amount', 'tdr_account_no', 'tdr_date', 'year']
    
    wb = load_workbook(path, read_only=True)
    correct = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(ws.cell(1, c).value or '') for c in range(1, 69)]
        if all(h.lower() == e.lower() for h, e in zip(headers, CORRECT)):
            correct += 1
            print(f"  ✅ {sheet_name}")
        else:
            print(f"  ❌ {sheet_name}")
    
    wb.close()
    
    print(f"\n✅ Complete! {correct}/{len(wb.sheetnames)} sheets verified")

except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
