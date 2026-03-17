#!/usr/bin/env python3
"""Update ALL sheets in Excel file with correct headers"""
import openpyxl

file_path = r'E:\kadhi\PMEGP_project\district_reports\KVIB_TG_district_details_01APR2022_to_31MAR2023_FIXED.xlsx'

correct_headers = [
    "id", "current_status", "under_process_agency_reason", "office_name", "agency_type", "state",
    "applicant_id", "applicant_name", "applicant_address", "applicant_mobile_no", "alternate_mobile_no",
    "email", "aadhar_no", "legal_status", "gender", "category", "special_category",
    "qualification", "date_of_birth", "age", "unit_location", "unit_address",
    "taluk_block", "unit_district", "industry_type", "product_desc_activity",
    "proposed_project_cost", "mm_involve", "financing_branch_ifsc_code", "financing_branch_address",
    "online_submission_date", "dltfec_meeting", "dltfec_meeting_place", "forwarding_date_to_bank",
    "bank_remarks", "date_of_documents_receiveda_at_bank",
    "project_cost_approved_ce", "project_cost_approved_wc", "project_cost_approved_total",
    "sanctioned_by_bank_date", "sanctioned_by_bank_ce", "sanctioned_by_bank_wc", "sanctioned_by_bank_total",
    "date_of_deposit_own_contribution", "own_contribution_amount_deposited",
    "covered_under_cgtsi", "date_of_loan_release", "loan_release_amount",
    "mm_claim_date", "mm_claim_amount", "remarks_for_mm_process_at_pmegp_co_mumbai",
    "mm_release_date", "mm_release_amount", "payment_status", "mm_disbursement_transaction_id", "fail_reason",
    "edp_training_center_name", "training_start_date", "training_end_date", "training_duration_days",
    "certificate_issue_date", "physical_verification_conducted_date", "physical_verification_status",
    "mm_final_adjustment_date", "mm_final_adjustment_amount",
    "tdr_account_no", "tdr_date", "year"
]

try:
    print(f"Opening: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    print(f"\nFound {len(wb.sheetnames)} sheet(s):")
    print(f"  {wb.sheetnames}")
    print()
    
    # Update each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Updating sheet: '{sheet_name}'")
        print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        # Update all 68 headers in row 1
        for col_idx, header in enumerate(correct_headers, 1):
            ws.cell(1, col_idx).value = header
        
        print(f"  ✓ Headers updated")
    
    # Save file
    print(f"\nSaving file...")
    wb.save(file_path)
    print(f"✅ SUCCESS!")
    print(f"All {len(wb.sheetnames)} sheet(s) updated with 68 correct database column headers")
    
except Exception as e:
    print(f"❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
