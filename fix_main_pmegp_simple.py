import openpyxl
from openpyxl.utils import get_column_letter
import os
import shutil
import sys

# Redirect output
output_file = open('fix_output.txt', 'w')

def log(msg):
    print(msg)
    output_file.write(msg + '\n')
    output_file.flush()

try:
    log("=" * 80)
    log("FIXING MAIN PMEGP FILE - SETTING CORRECT COLUMNS")
    log("=" * 80)
    
    # The 18 columns expected by the backend
    required_columns = [
        'rowNo', 'name', 'agencyReceived', 'agencyReturned', 'Pending_At_Agency',
        'Forwarded_to_Bank', 'sanctionedPrj', 'sanctionedLakh', 'claimedPrj',
        'claimedLakh', 'disbursementPrj', 'disbursementLakh', 'bankReturned',
        'pendingBankPrj', 'pendingBankLakh', 'pendingDisbursementPrj', 'pendingDisbursementLakh', 'year'
    ]
    
    input_file = r"excel/pmegp_TG_KVIB_district_pipeline_01APR2022_to_31MAR2023.xlsx"
    output_xlsx = r"excel/pmegp_TG_KVIB_district_pipeline_01APR2022_to_31MAR2023_FIXED.xlsx"
    
    log(f"\nInput file: {os.path.basename(input_file)}")
    log(f"Output file: {os.path.basename(output_xlsx)}")
    
    if not os.path.exists(input_file):
        log(f"\n✗ ERROR: Input file not found: {input_file}")
    else:
        log(f"✓ Input file found")
        
        # Create backup
        backup_file = input_file.replace('.xlsx', '_BACKUP.xlsx')
        if os.path.exists(backup_file):
            os.remove(backup_file)
        shutil.copy(input_file, backup_file)
        log(f"✓ Backup created: {os.path.basename(backup_file)}")
        
        # Load
        log("\n1. Loading workbook...")
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        log(f"   Sheet: {ws.title}")
        log(f"   Rows: {ws.max_row}, Cols: {ws.max_column}")
        
        # Preserve data
        log("\n2. Preserving data...")
        all_data = []
        for r in range(1, ws.max_row + 1):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 25)):
                cell_val = ws.cell(row=r, column=c).value
                row_data.append(cell_val)
            all_data.append(row_data)
        log(f"   ✓ Loaded {len(all_data)} rows")
        
        # Create new
        log("\n3. Creating fixed workbook...")
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        
        # Headers row 1
        log("   Setting headers in row 1:")
        for idx, hdr in enumerate(required_columns, 1):
            ws_new.cell(row=1, column=idx).value = hdr
        log(f"   ✓ Set 18 headers")
        
        # Keep row 2 if it exists
        if len(all_data) > 1:
            log("   Keeping row 2...")
            for c in range(1, min(19, len(all_data[1]) + 1)):
                ws_new.cell(row=2, column=c).value = all_data[1][c - 1]
        
        # Copy data rows
        log("   Copying data rows (row 3+)...")
        data_row = 3
        if len(all_data) > 2:
            for src_idx in range(2, len(all_data)):
                for c in range(1, min(19, len(all_data[src_idx]) + 1)):
                    ws_new.cell(row=data_row, column=c).value = all_data[src_idx][c - 1]
                data_row += 1
        log(f"   ✓ Copied {data_row - 3} rows")
        
        # Save
        log(f"\n4. Saving fixed file...")
        wb_new.save(output_xlsx)
        log(f"   ✓ Saved")
        
        log(f"\n{'='*80}")
        log(f"SUCCESS! Excel file fixed")
        log(f"{'='*80}")
        log(f"\nOutput: {output_xlsx}")
        log(f"Ready for upload")
        
except Exception as e:
    log(f"\n✗ Error: {e}")
    import traceback
    log(traceback.format_exc())

finally:
    output_file.close()
